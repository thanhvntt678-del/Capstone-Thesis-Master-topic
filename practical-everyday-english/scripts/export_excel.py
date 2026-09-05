#!/usr/bin/env python3
"""Export all lesson_*.json files into the Python-ready master Excel.

Column schema (fixed — Rule 6 / Rule 18, Python must never have to guess):
Lesson_ID, Major_Topic, Situation_ID, Situation_Title_EN, Situation_Title_VI,
CEFR, Turn_Order, Speaker, English, Vietnamese_Translation, Is_Learning_Text,
MP4_Part
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
LESSONS_DIR = ROOT / "data" / "lessons"
OUT_PATH = ROOT / "export" / "PRACTICAL_EVERYDAY_ENGLISH_MASTER.xlsx"

COLUMNS = [
    "Lesson_ID", "Major_Topic", "Situation_ID", "Situation_Title_EN",
    "Situation_Title_VI", "CEFR", "Turn_Order", "Speaker", "English",
    "Vietnamese_Translation", "Is_Learning_Text", "MP4_Part",
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DIALOGUE_LINES"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    lesson_files = sorted(LESSONS_DIR.glob("lesson_*.json"))
    row_idx = 2
    for path in lesson_files:
        data = json.loads(path.read_text(encoding="utf-8"))
        for turn in data.get("turns", []):
            row = [
                data["lesson_id"],
                data["major_topic"],
                data["situation_id"],
                data["situation_title_en"],
                data["situation_title_vi"],
                data["cefr"],
                turn["turn_order"],
                turn["speaker"],
                turn["english"],
                turn["vietnamese"],
                "TRUE" if turn.get("is_learning_text") else "FALSE",
                turn.get("mp4_part", 1),
            ]
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

    widths = [10, 26, 12, 34, 34, 10, 10, 12, 46, 46, 14, 10]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    ws.freeze_panes = "A2"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {row_idx - 2} dialogue rows from {len(lesson_files)} lessons to {OUT_PATH}")


if __name__ == "__main__":
    main()
